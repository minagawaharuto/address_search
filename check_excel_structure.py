from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
import sys

# UTF-8で出力
sys.stdout.reconfigure(encoding='utf-8')

# openpyxlのphonetic属性エラーを回避するためのパッチ
_original_column_dimension_init = ColumnDimension.__init__

def _patched_column_dimension_init(self, worksheet, *args, **kwargs):
    if 'phonetic' in kwargs:
        del kwargs['phonetic']
    _original_column_dimension_init(self, worksheet, *args, **kwargs)

ColumnDimension.__init__ = _patched_column_dimension_init

file_path = 'uploads/202512.xlsx'

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    print(f"ファイル: {file_path}")
    print("=" * 80)
    print(f"シート名: {ws.title}")
    print(f"最大行数: {ws.max_row}")
    print(f"最大列数: {ws.max_column}")

    print("\n" + "=" * 80)
    print("最初の20行のデータ（最初の10列のみ）:")
    print("=" * 80)

    for row in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(11, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value)[:30]
                row_data.append(f"[{col}]{cell_str}")

        if row_data:
            print(f"行{row}: {' | '.join(row_data)}")

    # ヘッダー行を探す
    print("\n" + "=" * 80)
    print("ヘッダー行の探索:")
    print("=" * 80)

    header_found = False
    for row in range(1, min(51, ws.max_row + 1)):
        row_values = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value:
                value_str = str(value).strip()
                row_values.append(value_str)

        row_str = ' '.join(row_values).lower()
        if '業態' in row_str or '店舗' in row_str or '店名' in row_str or '支店' in row_str:
            print(f"\n行{row}がヘッダーの可能性があります:")
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                if value:
                    print(f"  列{col}: {value}")
            header_found = True

    if not header_found:
        print("\n業態名・店舗名を含むヘッダー行が見つかりませんでした")

except Exception as e:
    print(f"エラー: {e}")
    import traceback
    traceback.print_exc()
