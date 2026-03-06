"""
Excelファイルの形式を確認するデバッグツール
使い方: python check_file.py ファイル名.xlsx
"""
import sys
import os


def check_file(file_path):
    """ファイルの詳細情報を表示"""
    print("=" * 60)
    print("Excelファイル診断ツール")
    print("=" * 60)

    # ファイルの存在確認
    if not os.path.exists(file_path):
        print(f"❌ エラー: ファイルが見つかりません: {file_path}")
        return

    print(f"✓ ファイルパス: {file_path}")

    # ファイルサイズ
    file_size = os.path.getsize(file_path)
    print(f"✓ ファイルサイズ: {file_size:,} bytes ({file_size / 1024:.2f} KB)")

    if file_size == 0:
        print("❌ エラー: ファイルが空です")
        return

    # ファイルヘッダーの読み取り
    try:
        with open(file_path, 'rb') as f:
            header = f.read(16)

        print(f"\n📄 ファイルヘッダー (最初の16バイト):")
        print(f"   HEX: {header.hex(' ')}")
        print(f"   バイナリ: {header[:8]}")

        # ファイル形式の判定
        print(f"\n🔍 ファイル形式の判定:")

        if header[:2] == b'PK':
            print("   ✓ Excel 2007+ (.xlsx) 形式 - ZIP/OpenXML 形式")
            print("   → openpyxl で読み込み可能")

            # ZIPファイルとして検証
            try:
                import zipfile
                if zipfile.is_zipfile(file_path):
                    print("   ✓ 有効なZIPファイルです")
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        files = zip_ref.namelist()
                        print(f"   ✓ ZIP内のファイル数: {len(files)}")
                        # Excel特有のファイルをチェック
                        has_workbook = any('workbook.xml' in f for f in files)
                        has_sheet = any('sheet' in f.lower() for f in files)
                        if has_workbook and has_sheet:
                            print("   ✓ Excelファイルの構造を確認")
                        else:
                            print("   ⚠ Excel構造が不完全かもしれません")
                else:
                    print("   ❌ ZIPファイルとして読み込めません（破損している可能性）")
            except Exception as e:
                print(f"   ❌ ZIP検証エラー: {e}")

        elif header[:4] == b'\xD0\xCF\x11\xE0':
            print("   ✓ Excel 97-2003 (.xls) 形式 - OLE2/BIFF 形式")
            print("   → xlrd で読み込み可能（変換が必要）")

        elif header[:2] == b'\x09\x08' or header[:2] == b'\x09\x04':
            print("   ⚠ BIFF2/BIFF3 形式（非常に古いExcel形式）")
            print("   → サポートされていません")

        else:
            print(f"   ❌ 不明なファイル形式")
            print(f"   → Excelファイルではない可能性があります")

            # 他の一般的なファイル形式をチェック
            if header[:4] == b'%PDF':
                print("   → PDFファイルです")
            elif header[:2] == b'\xFF\xD8':
                print("   → JPEGファイルです")
            elif header[:4] == b'\x89PNG':
                print("   → PNGファイルです")

    except Exception as e:
        print(f"❌ ファイル読み込みエラー: {e}")
        return

    # openpyxl での読み込みテスト
    print(f"\n🧪 openpyxl での読み込みテスト:")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        print(f"   ✓ 読み込み成功")
        print(f"   ✓ シート数: {len(wb.sheetnames)}")
        print(f"   ✓ シート名: {', '.join(wb.sheetnames)}")

        ws = wb.active
        print(f"   ✓ アクティブシート: {ws.title}")
        print(f"   ✓ 最大行数: {ws.max_row}")
        print(f"   ✓ 最大列数: {ws.max_column}")

        # 最初の数行を表示
        if ws.max_row > 0:
            print(f"\n   📊 最初の3行のデータ:")
            for row_idx, row in enumerate(ws.iter_rows(max_row=3, values_only=True), 1):
                print(f"      {row_idx}: {row}")

        wb.close()
        print("\n✅ 結論: このファイルは正常に読み込めます")

    except Exception as e:
        print(f"   ❌ 読み込み失敗: {e}")
        print(f"\n   💡 対処法:")
        print(f"      1. Excelでファイルを開いて「名前を付けて保存」→ .xlsx 形式で保存")
        print(f"      2. ファイルが破損していないか確認")
        print(f"      3. 古い .xls 形式の場合は変換が必要")

    # pandas での読み込みテスト
    print(f"\n🧪 pandas での読み込みテスト:")
    try:
        import pandas as pd

        # まずopenpyxlで試す
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            print(f"   ✓ openpyxl engine で読み込み成功")
        except:
            # xlrdで試す
            try:
                df = pd.read_excel(file_path, engine='xlrd')
                print(f"   ✓ xlrd engine で読み込み成功（古い形式）")
            except Exception as e:
                print(f"   ❌ 読み込み失敗: {e}")
                return

        print(f"   ✓ データ形状: {df.shape[0]}行 x {df.shape[1]}列")
        print(f"   ✓ 列名: {list(df.columns)}")

        if not df.empty:
            print(f"\n   📊 最初の3行:")
            print(df.head(3).to_string(index=False))
        else:
            print(f"   ⚠ データが空です")

    except Exception as e:
        print(f"   ❌ pandas テスト失敗: {e}")

    print("\n" + "=" * 60)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python check_file.py <ファイルパス>")
        print("例: python check_file.py サンプル.xlsx")
        sys.exit(1)

    file_path = sys.argv[1]
    check_file(file_path)
