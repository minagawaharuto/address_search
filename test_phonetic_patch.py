"""
openpyxlのphonetic属性パッチのテスト
"""
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
import os

# パッチを適用
_original_column_dimension_init = ColumnDimension.__init__

def _patched_column_dimension_init(self, worksheet, *args, **kwargs):
    if 'phonetic' in kwargs:
        print(f"[PATCH] phonetic属性を削除しました")
        del kwargs['phonetic']
    _original_column_dimension_init(self, worksheet, *args, **kwargs)

ColumnDimension.__init__ = _patched_column_dimension_init

print("=" * 70)
print("openpyxl phonetic属性パッチテスト")
print("=" * 70)

# テストファイルパス
test_file = os.path.join('uploads', '202512.xlsx')

if os.path.exists(test_file):
    print(f"\nテストファイル: {test_file}")
    print(f"ファイルサイズ: {os.path.getsize(test_file)} bytes")

    try:
        print("\nExcelファイル読み込み中...")
        wb = load_workbook(test_file, data_only=False)
        ws = wb.active

        print(f"✓ 読み込み成功!")
        print(f"  シート名: {ws.title}")
        print(f"  行数: {ws.max_row}")
        print(f"  列数: {ws.max_column}")

        # 最初の数行を表示
        print(f"\n最初の3行:")
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value)[:20] if cell_value else '')
            print(f"  {row_idx}: {' | '.join(row_data)}")

        wb.close()
        print("\n" + "=" * 70)
        print("✓ パッチが正常に動作しています!")
        print("=" * 70)

    except Exception as e:
        print(f"\n✗ エラー: {e}")
        import traceback
        traceback.print_exc()

else:
    print(f"\n✗ テストファイルが見つかりません: {test_file}")
    print("uploads/202512.xlsx を配置してからテストしてください")
