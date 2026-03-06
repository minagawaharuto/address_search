"""
サンプルExcelファイルを作成するスクリプト
使い方: python sample_branches.py
"""
from openpyxl import Workbook

# 新しいワークブックを作成
wb = Workbook()
ws = wb.active
ws.title = "支店リスト"

# ヘッダーを追加
ws['A1'] = '支店名'

# サンプルデータを追加
sample_branches = [
    '東京支店',
    '大阪支店',
    '名古屋支店',
    '福岡支店',
    '札幌支店',
]

for idx, branch in enumerate(sample_branches, start=2):
    ws[f'A{idx}'] = branch

# ファイルを保存
filename = 'サンプル_支店リスト.xlsx'
wb.save(filename)
print(f'サンプルファイル "{filename}" を作成しました')
print(f'支店数: {len(sample_branches)}件')
