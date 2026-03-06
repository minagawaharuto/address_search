"""
Excelファイルを診断して修正するツール
使い方: python fix_excel.py 元のファイル.xlsx
"""
import sys
import os
import pandas as pd
from openpyxl import Workbook


def fix_excel_file(input_file):
    """Excelファイルを読み込んで新しい形式で保存"""
    print("=" * 60)
    print("Excel ファイル修正ツール")
    print("=" * 60)
    print()

    # ファイルの存在確認
    if not os.path.exists(input_file):
        print(f"❌ エラー: ファイルが見つかりません: {input_file}")
        return False

    print(f"📁 入力ファイル: {input_file}")
    file_size = os.path.getsize(input_file)
    print(f"📊 ファイルサイズ: {file_size:,} bytes")
    print()

    # ファイル形式の検出
    with open(input_file, 'rb') as f:
        header = f.read(8)

    if header[:2] == b'PK':
        file_type = 'xlsx'
        print("✓ 形式: Excel 2007+ (.xlsx)")
    elif header[:4] == b'\xD0\xCF\x11\xE0':
        file_type = 'xls'
        print("✓ 形式: Excel 97-2003 (.xls)")
    else:
        print(f"❌ 不明な形式: {header.hex()}")
        return False

    print()
    print("🔄 ファイルを読み込んでいます...")

    # 複数の方法で読み込みを試す
    df = None
    read_method = None

    # 方法1: openpyxl
    try:
        df = pd.read_excel(input_file, engine='openpyxl')
        read_method = 'openpyxl'
        print("✓ openpyxl で読み込み成功")
    except Exception as e:
        print(f"⚠ openpyxl で失敗: {e}")

        # 方法2: xlrd
        if df is None:
            try:
                df = pd.read_excel(input_file, engine='xlrd')
                read_method = 'xlrd'
                print("✓ xlrd で読み込み成功")
            except Exception as e2:
                print(f"⚠ xlrd で失敗: {e2}")

        # 方法3: header=Noneで試す
        if df is None:
            try:
                df = pd.read_excel(input_file, engine='openpyxl', header=None)
                read_method = 'openpyxl (header=None)'
                print("✓ openpyxl (header=None) で読み込み成功")
            except Exception as e3:
                print(f"❌ すべての読み込み方法で失敗: {e3}")
                return False

    if df is None:
        print("❌ ファイルを読み込めませんでした")
        return False

    print()
    print("📊 データ情報:")
    print(f"   行数: {len(df)}")
    print(f"   列数: {len(df.columns)}")
    print(f"   列名: {list(df.columns)}")
    print()

    if len(df) > 0:
        print("📋 最初の5行:")
        print(df.head().to_string())
        print()

    # 新しいファイル名
    base_name = os.path.splitext(input_file)[0]
    output_file = f"{base_name}_修正済み.xlsx"

    print(f"💾 修正したファイルを保存しています...")

    try:
        # openpyxlで完全に新規作成
        wb = Workbook()
        ws = wb.active
        ws.title = "支店リスト"

        # ヘッダーを書き込み
        if df.columns[0] != 0:  # 列名がある場合
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=str(header))
            start_row = 2
        else:  # 列名がない場合
            ws.cell(row=1, column=1, value='支店名')
            start_row = 2

        # データを書き込み
        for row_idx, row_data in enumerate(df.values, start_row):
            for col_idx, value in enumerate(row_data, 1):
                if pd.notna(value):  # NaNでない場合のみ
                    ws.cell(row=row_idx, column=col_idx, value=str(value))

        wb.save(output_file)
        print(f"✅ 保存完了: {output_file}")
        print()

        # 保存したファイルのサイズ
        new_size = os.path.getsize(output_file)
        print(f"📊 新しいファイル:")
        print(f"   サイズ: {new_size:,} bytes")
        print(f"   場所: {os.path.abspath(output_file)}")
        print()

        # 検証
        print("🔍 保存したファイルを検証しています...")
        try:
            test_df = pd.read_excel(output_file, engine='openpyxl')
            print(f"✅ 検証成功: {len(test_df)}行読み込めました")
            print()
            print("=" * 60)
            print("✅ 修正完了！")
            print(f"新しいファイルをWebアプリでアップロードしてください:")
            print(f"  {output_file}")
            print("=" * 60)
            return True
        except Exception as e:
            print(f"❌ 検証失敗: {e}")
            return False

    except Exception as e:
        print(f"❌ 保存エラー: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python fix_excel.py <Excelファイルのパス>")
        print()
        print("例:")
        print('  python fix_excel.py "C:\\Users\\Panasonic\\Desktop\\支店リスト.xlsx"')
        print('  python fix_excel.py 支店リスト.xlsx')
        print()
        print("または、ファイルをこのフォルダにコピーして:")
        print('  python fix_excel.py ファイル名.xlsx')
        sys.exit(1)

    input_file = sys.argv[1]
    fix_excel_file(input_file)
