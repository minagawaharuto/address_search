# -*- coding: utf-8 -*-
"""
比較・修正機能のテスト
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# テスト用のExcelファイルを作成
def create_test_file():
    """テスト用のExcelファイルを作成"""
    wb = Workbook()
    ws = wb.active

    # ヘッダー行
    ws['A1'] = '業態名'
    ws['B1'] = '店舗名'
    ws['C1'] = '住所'
    ws['D1'] = '電話番号'
    ws['E1'] = '営業時間'

    # サンプルデータ（古い情報）
    ws['A2'] = 'ローソン'
    ws['B2'] = '東京駅前店'
    ws['C2'] = '東京都千代田区丸の内1-1-1'  # 古い住所
    ws['D2'] = '03-1234-5678'  # 古い電話番号
    ws['E2'] = '24時間営業'  # 古い営業時間

    ws['A3'] = 'セブンイレブン'
    ws['B3'] = '渋谷店'
    ws['C3'] = '東京都渋谷区渋谷1-1-1'
    ws['D3'] = '03-9876-5432'
    ws['E3'] = '24時間営業'

    # 保存
    test_file = 'test_comparison_input.xlsx'
    wb.save(test_file)
    print(f"テストファイルを作成しました: {test_file}")
    return test_file


# 比較関数のテスト
def test_normalize_and_compare():
    """正規化と比較関数のテスト"""
    from app_googlemaps import normalize_text, compare_data

    print("=" * 60)
    print("正規化と比較関数のテスト")
    print("=" * 60)

    # テストケース1: 完全一致
    text1 = "東京都千代田区丸の内1-1-1"
    text2 = "東京都千代田区丸の内1-1-1"
    result = compare_data(text1, text2)
    print(f"\nテスト1 (完全一致): {result}")
    print(f"  テキスト1: {text1}")
    print(f"  テキスト2: {text2}")
    print(f"  期待結果: False (差異なし)")
    print(f"  実際の結果: {result}")
    assert result == False, "完全一致のテストが失敗しました"

    # テストケース2: スペース違い
    text1 = "東京都 千代田区 丸の内1-1-1"
    text2 = "東京都千代田区丸の内1-1-1"
    result = compare_data(text1, text2)
    print(f"\nテスト2 (スペース違い): {result}")
    print(f"  テキスト1: {text1}")
    print(f"  テキスト2: {text2}")
    print(f"  期待結果: False (差異なし)")
    print(f"  実際の結果: {result}")
    assert result == False, "スペース違いのテストが失敗しました"

    # テストケース3: 内容が異なる
    text1 = "東京都千代田区丸の内1-1-1"
    text2 = "東京都港区六本木1-1-1"
    result = compare_data(text1, text2)
    print(f"\nテスト3 (内容が異なる): {result}")
    print(f"  テキスト1: {text1}")
    print(f"  テキスト2: {text2}")
    print(f"  期待結果: True (差異あり)")
    print(f"  実際の結果: {result}")
    assert result == True, "内容が異なるテストが失敗しました"

    # テストケース4: 片方が空
    text1 = "東京都千代田区丸の内1-1-1"
    text2 = ""
    result = compare_data(text1, text2)
    print(f"\nテスト4 (片方が空): {result}")
    print(f"  テキスト1: {text1}")
    print(f"  テキスト2: (空)")
    print(f"  期待結果: True (差異あり)")
    print(f"  実際の結果: {result}")
    assert result == True, "片方が空のテストが失敗しました"

    # テストケース5: 両方とも空
    text1 = ""
    text2 = ""
    result = compare_data(text1, text2)
    print(f"\nテスト5 (両方とも空): {result}")
    print(f"  テキスト1: (空)")
    print(f"  テキスト2: (空)")
    print(f"  期待結果: False (差異なし)")
    print(f"  実際の結果: {result}")
    assert result == False, "両方とも空のテストが失敗しました"

    print("\n" + "=" * 60)
    print("[OK] すべてのテストが成功しました！")
    print("=" * 60)


# ハイライト機能のテスト
def test_highlight():
    """ハイライト機能のテスト"""
    print("\n" + "=" * 60)
    print("ハイライト機能のテスト")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active

    # データを書き込む
    ws['A1'] = 'テストデータ'

    # 黄色ハイライトを適用
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    ws['A1'].fill = yellow_fill

    # 保存
    test_file = 'test_highlight.xlsx'
    wb.save(test_file)
    print(f"[OK] ハイライトテストファイルを作成しました: {test_file}")
    print("   Excelで開いて、セルA1が黄色でハイライトされていることを確認してください")


if __name__ == "__main__":
    try:
        # 比較関数のテスト
        test_normalize_and_compare()

        # ハイライト機能のテスト
        test_highlight()

        # テスト用Excelファイルを作成
        create_test_file()

        print("\n" + "=" * 60)
        print("[OK] すべてのテストが完了しました")
        print("=" * 60)
        print("\n次のステップ:")
        print("1. Webアプリを起動: python app_googlemaps.py")
        print("2. test_comparison_input.xlsx をアップロード")
        print("3. 処理後、黄色でハイライトされたセルを確認")

    except Exception as e:
        print(f"\n[ERROR] エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
