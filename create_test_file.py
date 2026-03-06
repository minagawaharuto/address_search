"""
テスト用のExcelファイルを作成
実行後、test_file.xlsx が作成されます
"""
from openpyxl import Workbook
import os

try:
    # 新しいワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = "支店リスト"

    # ヘッダーを追加
    ws['A1'] = '支店名'

    # サンプルデータを追加
    sample_branches = [
        '東京支店',
        '大阪支店',
        '名古屋支店',
    ]

    for idx, branch in enumerate(sample_branches, start=2):
        ws[f'A{idx}'] = branch

    # ファイルを保存
    filename = 'test_file.xlsx'
    wb.save(filename)

    file_path = os.path.abspath(filename)
    file_size = os.path.getsize(filename)

    print('✓ テストファイル作成成功！')
    print(f'  ファイル名: {filename}')
    print(f'  保存場所: {file_path}')
    print(f'  ファイルサイズ: {file_size} bytes')
    print(f'  データ行数: {len(sample_branches)}件')
    print('')
    print('このファイルをWebアプリでテストしてください。')
    print('会社名の例: セブンイレブン')

except Exception as e:
    print(f'❌ エラー: {e}')
    import traceback
    traceback.print_exc()
