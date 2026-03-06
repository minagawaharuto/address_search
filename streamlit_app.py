import streamlit as st
import os
import time
import re
import difflib
import pandas as pd
import requests
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import PatternFill, Alignment
from werkzeug.utils import secure_filename
import tempfile
import shutil

# openpyxl patch for phonetic attribute error
_original_column_dimension_init = ColumnDimension.__init__

def _patched_column_dimension_init(self, worksheet, *args, **kwargs):
    if 'phonetic' in kwargs:
        del kwargs['phonetic']
    _original_column_dimension_init(self, worksheet, *args, **kwargs)

ColumnDimension.__init__ = _patched_column_dimension_init

# Logging setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Helper Functions ---

def load_api_key():
    """Load API Key from st.secrets, env var, or config.txt"""
    # 1. Try st.secrets (for Streamlit Cloud)
    try:
        if "GOOGLE_MAPS_API_KEY" in st.secrets:
            return st.secrets["GOOGLE_MAPS_API_KEY"]
    except FileNotFoundError:
        pass # Not running on Streamlit Cloud or secrets.toml not found

    # 2. Try Environment Variable
    if "GOOGLE_MAPS_API_KEY" in os.environ:
        return os.environ["GOOGLE_MAPS_API_KEY"]

    # 3. Try config.txt (Local development)
    config_file = 'config.txt'
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('GOOGLE_MAPS_API_KEY='):
                        return line.split('=', 1)[1].strip()
        except Exception:
            pass
    return None

def detect_file_type(file_path):
    """Detect actual file type"""
    try:
        with open(file_path, 'rb') as f:
            header = f.read(8)

        if header[:2] == b'PK':
            return 'xlsx', 'Excel 2007+ (.xlsx) - ZIP format'
        elif header[:4] == b'\xD0\xCF\x11\xE0':
            return 'xls', 'Excel 97-2003 (.xls) - OLE2 format'
        else:
            return 'unknown', f'Unknown format (Header: {header.hex()})'
    except Exception as e:
        return 'error', str(e)

def convert_to_xlsx(input_file, output_file):
    """Convert old Excel format to .xlsx"""
    try:
        file_type, type_desc = detect_file_type(input_file)
        
        if file_type == 'unknown':
            return False, f"Unsupported file format: {type_desc}"

        try:
            if file_type == 'xls':
                df = pd.read_excel(input_file, engine='xlrd')
            else:
                df = pd.read_excel(input_file, engine='openpyxl')

            if df.empty:
                return False, "Excel file is empty"

            df.to_excel(output_file, index=False, engine='openpyxl')
            return True, "File converted successfully"

        except Exception as e:
            return False, f"Read failed: {str(e)}"

    except Exception as e:
        return False, f"Conversion error: {str(e)}"

def generate_search_patterns_simple(search_keyword):
    """Generate search patterns"""
    patterns = []
    keyword_clean = str(search_keyword).strip()
    patterns.append(keyword_clean)
    has_suffix = any(suffix in keyword_clean for suffix in ['支店', '店舗', '店', '営業所', '出張所', 'センター'])
    if not has_suffix:
        patterns.extend([f"{keyword_clean}店", f"{keyword_clean}支店"])
    return patterns

def search_address_with_googlemaps(search_keyword, api_key):
    """Search using Places API (New)"""
    try:
        if not api_key:
            return {'status': 'error', 'message': 'API Key not set'}

        url = "https://places.googleapis.com/v1/places:searchText"
        search_patterns = generate_search_patterns_simple(search_keyword)
        
        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": (
                "places.displayName,"
                "places.formattedAddress,"
                "places.addressComponents,"
                "places.nationalPhoneNumber,"
                "places.internationalPhoneNumber,"
                "places.location,"
                "places.regularOpeningHours,"
                "places.websiteUri,"
                "places.rating,"
                "places.userRatingCount,"
                "places.businessStatus,"
                "places.id"
            )
        }

        for i, search_query in enumerate(search_patterns):
            data = {"textQuery": search_query, "languageCode": "ja"}
            response = requests.post(url, json=data, headers=headers, timeout=10)

            if response.status_code == 200:
                result = response.json()
                if 'places' in result and len(result['places']) > 0:
                    place = result['places'][0]
                    
                    # 住所の整形（日本、郵便番号を削除）
                    raw_address = place.get('formattedAddress', '')
                    raw_address = re.sub(r'^日本、', '', raw_address)
                    raw_address = re.sub(r'〒\d{3}-\d{4}\s*', '', raw_address)
                    
                    place_info = {
                        'status': 'success',
                        'building_name': place.get('displayName', {}).get('text', ''),
                        'address': raw_address.strip(),
                        'postal_code': '',
                        'phone': place.get('nationalPhoneNumber', ''),
                        'latitude': '',
                        'longitude': '',
                        'business_hours': '',
                        'open_now': '',
                        'website': place.get('websiteUri', ''),
                        'rating': str(place.get('rating', '')),
                        'review_count': str(place.get('userRatingCount', '')),
                        'business_status': ''
                    }

                    if 'addressComponents' in place:
                        for component in place['addressComponents']:
                            if 'postal_code' in component.get('types', []):
                                place_info['postal_code'] = component.get('longText', '')
                                break
                    
                    if 'location' in place:
                        place_info['latitude'] = str(place['location'].get('latitude', ''))
                        place_info['longitude'] = str(place['location'].get('longitude', ''))

                    if 'regularOpeningHours' in place:
                        opening_hours = place['regularOpeningHours']
                        if 'weekdayDescriptions' in opening_hours:
                            place_info['business_hours'] = ' ・ '.join(opening_hours['weekdayDescriptions'])
                        if 'openNow' in opening_hours:
                            place_info['open_now'] = '営業中' if opening_hours['openNow'] else '営業時間外'

                    if 'businessStatus' in place:
                        status_map = {'OPERATIONAL': '営業中', 'CLOSED_TEMPORARILY': '一時休業', 'CLOSED_PERMANENTLY': '閉業'}
                        place_info['business_status'] = status_map.get(place['businessStatus'], place['businessStatus'])

                    return place_info

            if i < len(search_patterns) - 1:
                time.sleep(0.3)

        return {'status': 'not_found', 'message': 'Not found'}

    except Exception as e:
        return {'status': 'error', 'message': f'Error: {str(e)}'}

def search_single_row_excel(idx, business_type, store_name, total, api_key):
    """Search for a single row"""
    try:
        if business_type and store_name:
            search_query = f"{business_type} {store_name}"
        elif store_name:
            search_query = str(store_name)
        elif business_type:
            search_query = str(business_type)
        else:
            return idx, None, False

        result = search_address_with_googlemaps(search_query, api_key)

        if result['status'] == 'success':
            return idx, {
                'building_name': result.get('building_name', ''),
                'address': result.get('address', ''),
                'phone': result.get('phone', ''),
                'business_hours': result.get('business_hours', '')
            }, True
        else:
            return idx, {
                'building_name': result.get('message', 'Not found'),
                'address': '',
                'phone': '',
                'business_hours': ''
            }, False
    except Exception as e:
        return idx, {'building_name': f"Error: {str(e)}", 'address':'', 'phone':'', 'business_hours':''}, False


def normalize_address(address):
    """Normalize address for robust comparison"""
    if not address or pd.isna(address):
        return ""
    
    address = str(address).strip()
    
    # 1. Remove Japan and Postal Code
    address = re.sub(r'^日本、', '', address)
    address = re.sub(r'〒\d{3}-\d{4}\s*', '', address)
    
    # 2. Convert Full-width numbers/hyphens to Half-width (Simple char map for common ones)
    trans_map = str.maketrans({
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
        '－': '-', 'ー': '-', '−': '-'
    })
    address = address.translate(trans_map)
    
    # 3. Convert Kanji numbers to Arabic (Simple 0-9 mapping for addresses)
    # Note: Complex Kanji numbers like '二十' -> 20 are rare in address suffixes like Chome, usually it's '一丁目'
    kanji_map = str.maketrans({
        '〇': '0', '一': '1', '二': '2', '三': '3', '四': '4',
        '五': '5', '六': '6', '七': '7', '八': '8', '九': '9'
    })
    address = address.translate(kanji_map)
    
    # 4. Unify delimiters (Chome, Ban, Go, No -> -)
    address = re.sub(r'(丁目|番地|番|号|の)', '-', address)
    
    # 5. Remove spaces and cleanup multiple hyphens
    address = address.replace(' ', '').replace('　', '')
    address = re.sub(r'-+', '-', address) # 1--2 -> 1-2
    address = address.strip('-') # Remove leading/trailing hyphens
    
    return address

def is_address_different(addr1, addr2, threshold=0.85):
    """Check if addresses are effectively different using fuzzy matching"""
    norm1 = normalize_address(addr1)
    norm2 = normalize_address(addr2)
    
    if not norm1 and not norm2:
        return False
    if not norm1 and norm2:
        return True # New value added
    if norm1 and not norm2:
        return True # Value removed (unlikely from API but possible)
        
    # If normalized strings are exactly same
    if norm1 == norm2:
        return False
        
    # Calculate similarity ratio
    ratio = difflib.SequenceMatcher(None, norm1, norm2).ratio()
    
    # If similarity is high, consider them SAME (return False for is_different)
    return ratio < threshold

def normalize_phone(phone):
    """電話番号を正規化"""
    if not phone or pd.isna(phone):
        return ""
    
    phone = str(phone).strip()
    
    # 全角数字・ハイフンを半角に
    trans_map = str.maketrans({
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
        '－': '-', 'ー': '-', '−': '-'
    })
    phone = phone.translate(trans_map)
    
    # 記号削除
    phone = re.sub(r'[ \-\(\)]', '', phone)
    
    # 国際番号対応 (+81)
    if phone.startswith('+81'):
        phone = '0' + phone[3:]
        
    return phone

def is_phone_different(val1, val2):
    """電話番号比較"""
    norm1 = normalize_phone(val1)
    norm2 = normalize_phone(val2)
    
    if not norm1 and not norm2:
        return False
    if not norm1 and norm2:
        return True
        
    return norm1 != norm2


def process_excel_streamlit(file_path, output_path, api_key, progress_bar, status_text, max_workers=10, filter_business_type=None):
# ... (rest of the function, keeping imports and normalize_address separate)
    """Process Excel file with Streamlit feedback"""
    try:
        # Load workbook to find header
        temp_wb = load_workbook(file_path, data_only=True)
        temp_ws = temp_wb.active

        header_row_idx = 0
        for row in range(1, min(51, temp_ws.max_row + 1)):
            row_values = []
            for col in range(1, temp_ws.max_column + 1):
                value = temp_ws.cell(row=row, column=col).value
                if value:
                    row_values.append(str(value).strip().lower())
            row_str = ' '.join(row_values)
            if '業態' in row_str and ('店舗' in row_str or '店名' in row_str):
                header_row_idx = row - 1
                break

        df = pd.read_excel(file_path, engine='openpyxl', header=header_row_idx)

        # Identify columns
        business_type_col_idx = None
        store_name_col_idx = None
        original_address_col_idx = None
        original_phone_col_idx = None
        
        for idx, col_name in enumerate(df.columns):
            col_str = str(col_name).strip()
            if col_str == '業態名' or col_str == '業態':
                business_type_col_idx = idx
            elif '業態' in col_str and business_type_col_idx is None:
                business_type_col_idx = idx
            
            if col_str in ['店舗名', '店名', '支店名', '店舗', '支店']:
                store_name_col_idx = idx
            
            # Try to identify original address column
            if '住所' in col_str or 'address' in col_str.lower():
                if original_address_col_idx is None:
                    original_address_col_idx = idx
                elif col_str == '住所':
                    original_address_col_idx = idx
                    
            # Try to identify original phone column
            if '電話' in col_str or 'tel' in col_str or 'phone' in col_str:
                if original_phone_col_idx is None:
                    original_phone_col_idx = idx
                elif '電話' in col_str:
                    original_phone_col_idx = idx

        
        if store_name_col_idx is None:
            for idx, col_name in enumerate(df.columns):
                col_str = str(col_name).strip()
                if '清掃' in col_str or '時間' in col_str or '日' in col_str or 'No' in col_str:
                    continue
                if '店舗' in col_str or '店名' in col_str or '支店' in col_str:
                    store_name_col_idx = idx
                    break

        if business_type_col_idx is None and store_name_col_idx is None:
            store_name_col_idx = 0
            st.warning("業態名・店舗名の列が見つかりません。1列目を検索キーワードとして使用します。")

        # Setup new columns
        new_columns = []
        target_columns = {}
        original_col_map = {}
        current_col = 1
        
        for i, col_name in enumerate(df.columns):
            col_str = str(col_name).strip()
            col_lower = col_str.lower()
            
            # Skip existing result columns to re-insert them at correct positions
            if col_str in ['住所(検索結果)', '電話番号検索結果', '電話番号(検索結果)']:
                continue

            new_columns.append(col_name)
            original_col_map[i] = current_col
            
            is_address_col = False
            is_phone_col = False
            
            # Map existing columns to targets
            if col_str == '住所':
                target_columns['address_orig'] = current_col
                is_address_col = True
            elif ('住所' in col_lower or 'address' in col_lower) and 'address_orig' not in target_columns:
                target_columns['address_orig'] = current_col
                is_address_col = True
            
            if col_str == '電話番号' or col_str == '電話':
                target_columns['phone_orig'] = current_col
                is_phone_col = True
            elif ('電話' in col_lower or 'phone' in col_lower) and 'phone_orig' not in target_columns:
                target_columns['phone_orig'] = current_col
                is_phone_col = True
                
            if ('営業時間' in col_lower or '営業' in col_lower) and 'business_hours' not in target_columns:
                target_columns['business_hours'] = current_col

            current_col += 1
            
            # Insert result columns immediately after originals
            if is_address_col and 'address_result' not in target_columns:
                new_columns.append('住所(検索結果)')
                target_columns['address_result'] = current_col
                current_col += 1
                
            if is_phone_col and 'phone_result' not in target_columns:
                new_columns.append('電話番号検索結果')
                target_columns['phone_result'] = current_col
                current_col += 1

        # Add missing columns
        if 'address_orig' not in target_columns:
            new_columns.append('住所')
            target_columns['address_orig'] = current_col
            current_col += 1
            
            new_columns.append('住所(検索結果)')
            target_columns['address_result'] = current_col
            current_col += 1

        elif 'address_result' not in target_columns:
            new_columns.append('住所(検索結果)')
            target_columns['address_result'] = current_col
            current_col += 1
            
        if 'phone_orig' not in target_columns:
            new_columns.append('電話番号')
            target_columns['phone_orig'] = current_col
            current_col += 1
            
            new_columns.append('電話番号検索結果')
            target_columns['phone_result'] = current_col
            current_col += 1

        elif 'phone_result' not in target_columns:
            new_columns.append('電話番号検索結果')
            target_columns['phone_result'] = current_col
            current_col += 1
            
        if 'business_hours' not in target_columns:
            new_columns.append('営業時間')
            target_columns['business_hours'] = current_col
            current_col += 1
        
        # '正式名称' is always added as a new column for clarity
        new_columns.append('正式名称')
        target_columns['building_name'] = current_col
        current_col += 1

        wb = Workbook()
        ws = wb.active

        for col_idx, col_name in enumerate(new_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data):
                if col_idx in original_col_map:
                    ws.cell(row=row_idx, column=original_col_map[col_idx], value=value)

        business_type_col = original_col_map[business_type_col_idx] if business_type_col_idx is not None else None
        store_name_col = original_col_map[store_name_col_idx] if store_name_col_idx is not None else None
        
        tasks = []
        row_mapping = {}

        for row in range(2, ws.max_row + 1):
            if business_type_col and store_name_col:
                business_type = ws.cell(row=row, column=business_type_col).value
                store_name = ws.cell(row=row, column=store_name_col).value
            else:
                business_type = None
                store_name = ws.cell(row=row, column=store_name_col).value

            if not business_type and not store_name:
                continue

            if filter_business_type:
                if not business_type or filter_business_type.lower() not in str(business_type).strip().lower():
                    continue

            if (business_type or store_name) and str(store_name or business_type).strip() and str(store_name or business_type) != 'nan':
                idx = len(tasks)
                tasks.append((idx, business_type, store_name))
                row_mapping[idx] = row

        processed_count = 0
        success_count = 0
        total = len(tasks)

        # Fill style for highlighting
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        if total == 0:
            st.warning("処理対象のデータが見つかりませんでした。")
            return False, "No data to process"

        status_text.text(f"Starting processing for {total} items...")
        progress_bar.progress(0)

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(search_single_row_excel, idx, business_type, store_name, total, api_key): idx
                for idx, business_type, store_name in tasks
            }

            for future in as_completed(futures):
                try:
                    idx, data, is_success = future.result()
                    if data is not None:
                        row = row_mapping[idx]
                        
                        # 取得したデータを各列に書き込む
                        if 'building_name' in target_columns:
                            cell = ws.cell(row=row, column=target_columns['building_name'], value=data['building_name'])
                            cell.alignment = Alignment(wrap_text=True)
                        
                        # 住所の処理
                        if is_success:
                            # 検索結果を書き込む
                            if 'address_result' in target_columns:
                                cell_res = ws.cell(row=row, column=target_columns['address_result'], value=data['address'])
                                cell_res.alignment = Alignment(wrap_text=True)
                            
                            # 元の住所を訂正・ハイライト
                            if 'address_orig' in target_columns:
                                cell_orig = ws.cell(row=row, column=target_columns['address_orig'])
                                original_val = cell_orig.value
                                cell_orig.alignment = Alignment(wrap_text=True)
                                
                                # 比較して差異があれば訂正してハイライト
                                if is_address_different(original_val, data['address']):
                                    cell_orig.value = data['address']
                                    cell_orig.fill = yellow_fill
                        
                        # 電話番号の処理
                        if is_success:
                            # 検索結果を書き込む
                            if 'phone_result' in target_columns:
                                cell_res = ws.cell(row=row, column=target_columns['phone_result'], value=data['phone'])
                                cell_res.alignment = Alignment(wrap_text=True)
                            
                            # 元の電話番号を訂正・ハイライト
                            if 'phone_orig' in target_columns:
                                cell_orig = ws.cell(row=row, column=target_columns['phone_orig'])
                                original_val = cell_orig.value
                                cell_orig.alignment = Alignment(wrap_text=True)
                                
                                # 比較して差異があれば訂正してハイライト
                                if is_phone_different(original_val, data['phone']):
                                    cell_orig.value = data['phone']
                                    cell_orig.fill = yellow_fill

                        if 'business_hours' in target_columns:
                            cell = ws.cell(row=row, column=target_columns['business_hours'], value=data['business_hours'])
                            cell.alignment = Alignment(wrap_text=True)

                        processed_count += 1
                        if is_success:
                            success_count += 1
                        
                        progress = processed_count / total
                        progress_bar.progress(progress)
                        status_text.text(f"Processing... {processed_count}/{total} (Success: {success_count})")
                except Exception as e:
                    st.error(f"Error processing row: {e}")

        # Preserve styles (simplified)
        new_wb = Workbook()
        new_ws = new_wb.active
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
            for col_idx, cell in enumerate(row, 1):
                new_cell = new_ws.cell(row=row_idx, column=col_idx)
                new_cell.value = cell.value
                # Copy fill for highlighted cells
                if cell.fill and cell.fill.fill_type:
                    new_cell.fill = cell.fill.copy()

        new_wb.save(output_path)
        return True, f"処理完了: {processed_count}件中{success_count}件の情報を取得しました"

    except Exception as e:
        import traceback
        st.error(f"Processing error: {str(e)}\n{traceback.format_exc()}")
        return False, str(e)


# --- Streamlit UI ---

st.set_page_config(page_title="住所検索システム", page_icon="🗺️")

st.title("🗺️ 住所検索システム")
st.markdown("""
Excelファイルをアップロードして、Google Maps API（従量課金）を使用して住所、電話番号、営業時間などの情報を自動取得します。
""")

# API Key Handling
st.markdown("### 🔑 APIキーの設定")
st.markdown("""
このシステムを利用するには、**Google Maps Platform の APIキー**が必要です。
ご自身のAPIキーを入力してください。（**Places API (New)** が有効になっている必要があります）
""")
st.caption("⚠️ 入力されたAPIキーはこのブラウザセッションでのみ使用され、サーバーには保存されません。")

# ローカル設定やSecretsがあれば読み込むが、基本は空欄でもOK
default_api_key = load_api_key()
api_key = st.text_input("Google Maps API Key", value=default_api_key if default_api_key else "", type="password", placeholder="AIzaSy...")

if not api_key:
    st.warning("👆 上記にAPIキーを入力すると、ファイルのアップロードが可能になります。")

# File Upload
uploaded_file = st.file_uploader("Excelファイルをアップロード (.xlsx, .xls)", type=['xlsx', 'xls'])
filter_business_type = st.text_input("業態名フィルタ (オプション)", help="特定の業態のみを処理したい場合に入力してください")

if uploaded_file and api_key:
    if st.button("処理開始"):
        # Create temp directory
        with tempfile.TemporaryDirectory() as tmpdirname:
            # Save uploaded file
            input_path = os.path.join(tmpdirname, secure_filename(uploaded_file.name))
            with open(input_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Check/Convert file
            is_valid, result = validate_excel_file = True, None # Simplified check here, trusting process logic or re-implementing basic check
            
            file_type, type_desc = detect_file_type(input_path)
            st.info(f"検出されたファイル形式: {type_desc}")
            
            process_path = input_path
            if file_type == 'xls':
                converted_path = os.path.join(tmpdirname, "converted.xlsx")
                success, msg = convert_to_xlsx(input_path, converted_path)
                if success:
                    process_path = converted_path
                    st.success(msg)
                else:
                    st.error(msg)
                    st.stop()
            elif file_type == 'unknown':
                st.error("不明なファイル形式です")
                st.stop()
            
            output_filename = f"住所追加_{os.path.splitext(uploaded_file.name)[0]}.xlsx"
            output_path = os.path.join(tmpdirname, output_filename)
            
            # Progress UI
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Process
            success, message = process_excel_streamlit(
                process_path, 
                output_path, 
                api_key, 
                progress_bar, 
                status_text,
                filter_business_type=filter_business_type
            )
            
            if success:
                st.success(message)
                with open(output_path, "rb") as f:
                    st.download_button(
                        label="結果をダウンロード",
                        data=f,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.error(f"処理に失敗しました: {message}")
