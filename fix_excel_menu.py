"""
Excelファイルを選択して修正するツール（メニュー形式）
使い方: python fix_excel_menu.py
"""
import os
import glob
import pandas as pd
from openpyxl import Workbook


def fix_excel_file(input_file, output_file):
    """Excelファイルを読み込んで新しい形式で保存"""
    print()
    print("🔄 ファイルを読み込んでいます...")

    # 複数の方法で読み込みを試す
    df = None

    # 方法1: openpyxl
    try:
        df = pd.read_excel(input_file, engine='openpyxl')
        print("✓ openpyxl で読み込み成功")
    except Exception as e:
        print(f"⚠ openpyxl で失敗: {e}")

        # 方法2: xlrd
        try:
            df = pd.read_excel(input_file, engine='xlrd')
            print("✓ xlrd で読み込み成功")
        except Exception as e2:
            print(f"⚠ xlrd で失敗: {e2}")

            # 方法3: header=Noneで試す
            try:
                df = pd.read_excel(input_file, engine='openpyxl', header=None)
                print("✓ openpyxl (header=None) で読み込み成功")
            except Exception as e3:
                print(f"❌ すべての読み込み方法で失敗: {e3}")
                return False

    if df is None:
        print("❌ ファイルを読み込めませんでした")
        return False

    print()
    print("📊 データ情報:")
    print(f"   行数: {len(df)}")
    print(f"   列数: {len(df.columns)}")
    print()

    if len(df) > 0:
        print("📋 最初の3行:")
        print(df.head(3).to_string())
        print()

    print(f"💾 修正したファイルを保存しています...")

    try:
        # openpyxlで完全に新規作成
        wb = Workbook()
        ws = wb.active
        ws.title = "支店リスト"

        # ヘッダーを書き込み
        if df.columns[0] != 0:  # 列名がある場合
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=str(header))
            start_row = 2
        else:  # 列名がない場合
            ws.cell(row=1, column=1, value='支店名')
            start_row = 2

        # データを書き込み
        for row_idx, row_data in enumerate(df.values, start_row):
            for col_idx, value in enumerate(row_data, 1):
                if pd.notna(value):  # NaNでない場合のみ
                    ws.cell(row=row_idx, column=col_idx, value=str(value))

        wb.save(output_file)
        print(f"✅ 保存完了: {output_file}")
        print()

        # 検証
        print("🔍 保存したファイルを検証しています...")
        test_df = pd.read_excel(output_file, engine='openpyxl')
        print(f"✅ 検証成功: {len(test_df)}行読み込めました")
        return True

    except Exception as e:
        print(f"❌ 保存エラー: {e}")
        return False


def main():
    print("=" * 60)
    print("Excel ファイル修正ツール（メニュー版）")
    print("=" * 60)
    print()

    # 現在のフォルダ内の.xlsxファイルを検索
    current_dir = os.getcwd()
    xlsx_files = glob.glob("*.xlsx")
    xls_files = glob.glob("*.xls")
    all_files = xlsx_files + xls_files

    if not all_files:
        print("❌ このフォルダにExcelファイルが見つかりません")
        print(f"   現在のフォルダ: {current_dir}")
        print()
        print("対処方法:")
        print("1. Excelファイルをこのフォルダにコピーする")
        print("2. または、フルパスを指定:")
        print('   python fix_excel.py "C:\\path\\to\\file.xlsx"')
        return

    print(f"📁 現在のフォルダ: {current_dir}")
    print()
    print("見つかったExcelファイル:")
    print()

    for idx, file in enumerate(all_files, 1):
        size = os.path.getsize(file)
        print(f"  {idx}. {file} ({size:,} bytes)")

    print()
    print("0. キャンセル")
    print()

    # ファイルを選択
    while True:
        try:
            choice = input("修正するファイルの番号を入力してください: ").strip()

            if choice == "0":
                print("キャンセルしました")
                return

            choice_num = int(choice)

            if 1 <= choice_num <= len(all_files):
                selected_file = all_files[choice_num - 1]
                break
            else:
                print(f"❌ 1〜{len(all_files)}の番号を入力してください")

        except ValueError:
            print("❌ 数字を入力してください")
        except KeyboardInterrupt:
            print("\nキャンセルしました")
            return

    print()
    print("=" * 60)
    print(f"選択されたファイル: {selected_file}")
    print("=" * 60)

    # 出力ファイル名
    base_name = os.path.splitext(selected_file)[0]
    output_file = f"{base_name}_修正済み.xlsx"

    # 修正実行
    success = fix_excel_file(selected_file, output_file)

    if success:
        print()
        print("=" * 60)
        print("✅ 修正完了！")
        print()
        print("次のステップ:")
        print("1. Webアプリを起動: python app.py")
        print("2. ブラウザで http://localhost:5000 を開く")
        print("3. 以下のファイルをアップロード:")
        print(f"   {output_file}")
        print("=" * 60)
    else:
        print()
        print("=" * 60)
        print("❌ 修正に失敗しました")
        print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nキャンセルしました")
    except Exception as e:
        print(f"\n❌ エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
