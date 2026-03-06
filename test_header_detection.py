import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
import sys

# UTF-8で出力
sys.stdout.reconfigure(encoding='utf-8')

# openpyxlのphonetic属性エラーを回避するためのパッチ
_original_column_dimension_init = ColumnDimension.__init__

def _patched_column_dimension_init(self, worksheet, *args, **kwargs):
    if 'phonetic' in kwargs:
        del kwargs['phonetic']
    _original_column_dimension_init(self, worksheet, *args, **kwargs)

ColumnDimension.__init__ = _patched_column_dimension_init

file_path = 'uploads/202512.xlsx'

print("=" * 80)
print("ヘッダー行検出テスト")
print("=" * 80)

# まずopenpyxlでヘッダー行を検出
temp_wb = load_workbook(file_path, data_only=True)
temp_ws = temp_wb.active

# ヘッダー行を検出（業態名・店舗名を含む行を探す）
header_row_idx = 0  # 0-indexed for pandas
for row in range(1, min(51, temp_ws.max_row + 1)):
    row_values = []
    for col in range(1, temp_ws.max_column + 1):
        value = temp_ws.cell(row=row, column=col).value
        if value:
            row_values.append(str(value).strip().lower())

    row_str = ' '.join(row_values)
    if '業態' in row_str and ('店舗' in row_str or '店名' in row_str):
        header_row_idx = row - 1  # pandas is 0-indexed
        print(f"✓ ヘッダー行を検出: 行{row} (pandas index: {header_row_idx})")
        break

# pandasでデータを読み込む（検出したヘッダー行を使用）
df = pd.read_excel(file_path, engine='openpyxl', header=header_row_idx)
print(f"✓ pandas読み込み成功: {len(df)}行 x {len(df.columns)}列")
print(f"✓ 列名: {df.columns.tolist()[:10]}")

# 業態名と店舗名の列を検出
business_type_col = None
store_name_col = None

for idx, col_name in enumerate(df.columns, 1):
    col_str = str(col_name).lower()
    if '業態' in col_str:
        business_type_col = col_name
        print(f"✓ 業態名列を検出: 列{idx} ({col_name})")
    if '店舗' in col_str or '店名' in col_str:
        store_name_col = col_name
        print(f"✓ 店舗名列を検出: 列{idx} ({col_name})")

print("\n" + "=" * 80)
print("業態名のユニークな値（最初の10件）:")
print("=" * 80)

if business_type_col:
    unique_values = df[business_type_col].dropna().unique()
    for idx, val in enumerate(unique_values[:10], 1):
        print(f"  {idx}. {val}")
    print(f"\n総ユニーク数: {len(unique_values)}種類")

    print("\n" + "=" * 80)
    print("フィルタテスト:")
    print("=" * 80)

    # 「魚屋路」でフィルタテスト
    test_filter = "魚屋路"
    matched_df = df[df[business_type_col].str.strip().str.lower() == test_filter.lower()]
    print(f"フィルタ「{test_filter}」: {len(matched_df)}件マッチ")

    if len(matched_df) > 0:
        print("\nマッチした最初の3件:")
        for idx, (_, row) in enumerate(matched_df.head(3).iterrows(), 1):
            print(f"  {idx}. {row[business_type_col]} - {row[store_name_col] if store_name_col else 'N/A'}")
