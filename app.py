import os
import time
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import PatternFill, Alignment
import pandas as pd
import requests
import logging

# openpyxlのphonetic属性エラーを回避するためのパッチ
_original_column_dimension_init = ColumnDimension.__init__

def _patched_column_dimension_init(self, worksheet, *args, **kwargs):
    # phonetic属性を削除（openpyxl 3.1.0以降の互換性問題対策）
    if 'phonetic' in kwargs:
        del kwargs['phonetic']
    _original_column_dimension_init(self, worksheet, *args, **kwargs)

ColumnDimension.__init__ = _patched_column_dimension_init

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24).hex()
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# ディレクトリの作成
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# ログ設定
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Google Maps APIキー（後で設定ファイルから読み込む）
GOOGLE_MAPS_API_KEY = None

# 元のファイルの列名を保存する辞書
original_columns = {}


def load_api_key():
    """APIキーを読み込む"""
    global GOOGLE_MAPS_API_KEY

    config_file = 'config.txt'
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('GOOGLE_MAPS_API_KEY='):
                        GOOGLE_MAPS_API_KEY = line.split('=', 1)[1].strip()
                        logger.info("APIキーを読み込みました")
                        return True
        except Exception as e:
            logger.error(f"APIキー読み込みエラー: {e}")

    logger.warning("APIキーが設定されていません。config.txtを確認してください")
    return False

# アプリケーション起動時にAPIキーを読み込む
load_api_key()


def detect_file_type(file_path):
    """ファイルの実際の形式を検出"""
    try:
        with open(file_path, 'rb') as f:
            header = f.read(8)

        if header[:2] == b'PK':
            return 'xlsx', 'Excel 2007+ (.xlsx) - ZIP形式'
        elif header[:4] == b'\xD0\xCF\x11\xE0':
            return 'xls', 'Excel 97-2003 (.xls) - OLE2形式'
        else:
            return 'unknown', f'不明な形式 (ヘッダー: {header.hex()})'
    except Exception as e:
        logger.error(f"ファイル形式検出エラー: {str(e)}")
        return 'error', str(e)


def convert_to_xlsx(input_file, output_file):
    """古い形式のExcelファイルを.xlsx形式に変換"""
    try:
        logger.info(f"ファイル変換開始: {input_file}")

        file_type, type_desc = detect_file_type(input_file)
        logger.info(f"検出されたファイル形式: {type_desc}")

        if file_type == 'unknown':
            return False, f"サポートされていないファイル形式です。実際のファイル形式: {type_desc}"

        try:
            if file_type == 'xls':
                df = pd.read_excel(input_file, engine='xlrd')
            else:
                df = pd.read_excel(input_file, engine='openpyxl')

            if df.empty:
                return False, "Excelファイルにデータが含まれていません"

            logger.info(f"データ読み込み成功: {len(df)}行, {len(df.columns)}列")

            df.to_excel(output_file, index=False, engine='openpyxl')
            logger.info(f"ファイル変換完了: {output_file}")

            return True, "ファイル形式を変換しました"

        except Exception as e:
            logger.error(f"pandas読み込みエラー: {str(e)}")
            return False, f"Excelファイルの読み込みに失敗しました: {str(e)}"

    except Exception as e:
        logger.error(f"ファイル変換エラー: {str(e)}")
        return False, f"ファイル変換エラー: {str(e)}"


def validate_excel_file(file_path):
    """Excelファイルの妥当性をチェック"""
    try:
        logger.info(f"ファイル検証開始: {file_path}")

        if not os.path.exists(file_path):
            return False, "ファイルが見つかりません"

        file_size = os.path.getsize(file_path)
        logger.info(f"ファイルサイズ: {file_size} bytes")

        if file_size == 0:
            return False, "ファイルが空です"

        file_type, type_desc = detect_file_type(file_path)
        logger.info(f"ファイル形式: {type_desc}")

        if file_type == 'xlsx':
            try:
                wb = load_workbook(file_path, read_only=True)
                wb.close()
                logger.info("openpyxlでの読み込み成功")
                return True, "ファイル形式は正常です"
            except Exception as e:
                logger.warning(f"openpyxlで開けません: {str(e)}")

        if file_type == 'xls' or file_type == 'xlsx':
            logger.info("ファイル変換を試みます")
            base_name = os.path.splitext(file_path)[0]
            converted_path = f"{base_name}_converted.xlsx"

            success, message = convert_to_xlsx(file_path, converted_path)
            if success:
                logger.info(f"変換成功: {converted_path}")
                return True, converted_path
            else:
                logger.error(f"変換失敗: {message}")
                return False, message
        else:
            return False, f"サポートされていないファイル形式です: {type_desc}"

    except Exception as e:
        logger.error(f"ファイル検証エラー: {str(e)}")
        return False, f"ファイル検証エラー: {str(e)}"


def generate_search_patterns_simple(search_keyword):
    """
    検索キーワードから複数の検索パターンを生成
    例: "ローソン 新宿" → ["ローソン 新宿", "ローソン 新宿店", "ローソン 新宿支店"]
    """
    patterns = []
    keyword_clean = str(search_keyword).strip()

    # そのまま使用
    patterns.append(keyword_clean)

    # 既に「支店」「店」などが含まれているか確認
    has_suffix = any(suffix in keyword_clean for suffix in ['支店', '店舗', '店', '営業所', '出張所', 'センター'])

    if not has_suffix:
        # 接尾辞がない場合は、バリエーションを追加
        patterns.extend([
            f"{keyword_clean}店",
            f"{keyword_clean}支店",
        ])

    return patterns


def search_address_with_googlemaps(search_keyword):
    """Places API (New)を使って住所・電話番号などの詳細情報を検索"""
    try:
        if not GOOGLE_MAPS_API_KEY:
            return {
                'status': 'error',
                'message': 'APIキーが設定されていません'
            }

        # Places API (New)のエンドポイント
        url = "https://places.googleapis.com/v1/places:searchText"

        # 検索クエリのバリエーションを生成
        search_patterns = generate_search_patterns_simple(search_keyword)

        logger.info(f"Places API (New)で検索中: {search_keyword} ({len(search_patterns)}パターン)")

        # ヘッダー（Field Maskで取得する情報を指定）
        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": GOOGLE_MAPS_API_KEY,
            "X-Goog-FieldMask": (
                "places.displayName,"
                "places.formattedAddress,"
                "places.addressComponents,"
                "places.nationalPhoneNumber,"
                "places.internationalPhoneNumber,"
                "places.location,"
                "places.regularOpeningHours,"
                "places.websiteUri,"
                "places.rating,"
                "places.userRatingCount,"
                "places.businessStatus,"
                "places.id"
            )
        }

        # 複数の検索パターンを順番に試す
        for i, search_query in enumerate(search_patterns):
            logger.info(f"  パターン{i+1}: {search_query}")

            # リクエストボディ
            data = {
                "textQuery": search_query,
                "languageCode": "ja"
            }

            # POSTリクエストを送信
            response = requests.post(url, json=data, headers=headers, timeout=10)

            # レスポンスのチェック
            if response.status_code == 200:
                result = response.json()

                if 'places' in result and len(result['places']) > 0:
                    place = result['places'][0]

                    # 住所の整形（日本、郵便番号を削除）
                    raw_address = place.get('formattedAddress', '')
                    raw_address = re.sub(r'^日本、', '', raw_address)
                    raw_address = re.sub(r'〒\d{3}-\d{4}\s*', '', raw_address)

                    # 各種情報を抽出
                    place_info = {
                        'status': 'success',
                        'building_name': place.get('displayName', {}).get('text', ''),
                        'address': raw_address.strip(),
                        'postal_code': '',
                        'phone': place.get('nationalPhoneNumber', ''),
                        'latitude': '',
                        'longitude': '',
                        'business_hours': '',
                        'open_now': '',
                        'website': place.get('websiteUri', ''),
                        'rating': '',
                        'review_count': '',
                        'business_status': '',
                        'search_query': search_query
                    }

                    # 郵便番号を抽出
                    if 'addressComponents' in place:
                        for component in place['addressComponents']:
                            if 'postal_code' in component.get('types', []):
                                place_info['postal_code'] = component.get('longText', '')
                                break

                    # 緯度経度を抽出
                    if 'location' in place:
                        place_info['latitude'] = str(place['location'].get('latitude', ''))
                        place_info['longitude'] = str(place['location'].get('longitude', ''))

                    # 営業時間を抽出
                    if 'regularOpeningHours' in place:
                        opening_hours = place['regularOpeningHours']
                        if 'weekdayDescriptions' in opening_hours:
                            # 営業時間を中点区切りで結合
                            place_info['business_hours'] = ' ・ '.join(opening_hours['weekdayDescriptions'])
                        if 'openNow' in opening_hours:
                            place_info['open_now'] = '営業中' if opening_hours['openNow'] else '営業時間外'

                    # 評価とレビュー数を抽出
                    if 'rating' in place:
                        place_info['rating'] = str(place['rating'])
                    if 'userRatingCount' in place:
                        place_info['review_count'] = str(place['userRatingCount'])

                    # 営業状況を抽出
                    if 'businessStatus' in place:
                        status_map = {
                            'OPERATIONAL': '営業中',
                            'CLOSED_TEMPORARILY': '一時休業',
                            'CLOSED_PERMANENTLY': '閉業'
                        }
                        place_info['business_status'] = status_map.get(place['businessStatus'], place['businessStatus'])

                    logger.info(f"  → 情報取得成功: {place_info['building_name']}")
                    return place_info

            # 次のパターンを試す前に少し待機（API制限対策）
            if i < len(search_patterns) - 1:
                time.sleep(0.3)

        # すべてのパターンで見つからなかった場合
        logger.warning(f"検索結果なし: {search_keyword} (全{len(search_patterns)}パターン試行)")
        return {
            'status': 'not_found',
            'message': '情報が見つかりませんでした'
        }

    except requests.exceptions.Timeout:
        logger.error(f"タイムアウト: {search_keyword}")
        return {
            'status': 'error',
            'message': 'エラー: リクエストがタイムアウトしました'
        }

    except requests.exceptions.RequestException as e:
        logger.error(f"リクエストエラー: {str(e)}")
        return {
            'status': 'error',
            'message': f'エラー: {str(e)}'
        }

    except Exception as e:
        logger.error(f"検索エラー: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'status': 'error',
            'message': f'エラー: {str(e)}'
        }


def search_single_row_excel(idx, business_type, store_name, total):
    """単一の行を検索（Excel並列処理用）"""
    try:
        # 検索クエリを作成
        if business_type and store_name:
            search_query = f"{business_type} {store_name}"
        elif store_name:
            search_query = str(store_name)
        elif business_type:
            search_query = str(business_type)
        else:
            return idx, None, False

        logger.info(f"処理中: {idx + 1}/{total} - {search_query}")

        # 住所を検索
        result = search_address_with_googlemaps(search_query)

        # 結果を返す
        if result['status'] == 'success':
            return idx, {
                'building_name': result.get('building_name', ''),
                'address': result.get('address', ''),
                'phone': result.get('phone', ''),
                'business_hours': result.get('business_hours', '')
            }, True
        else:
            return idx, {
                'building_name': result.get('message', '情報が見つかりませんでした'),
                'address': '',
                'phone': '',
                'business_hours': ''
            }, False

    except Exception as e:
        logger.error(f"検索エラー (行{idx + 1}): {str(e)}")
        return idx, {
            'building_name': f"エラー: {str(e)}",
            'address': '',
            'phone': '',
            'business_hours': ''
        }, False


import difflib

def normalize_text(text):
    """テキストを正規化して比較しやすくする（強化版）"""
    if not text or pd.isna(text):
        return ""
    
    text = str(text).strip()
    
    # 1. 日本、郵便番号を削除
    text = re.sub(r'^日本、', '', text)
    text = re.sub(r'〒\d{3}-\d{4}\s*', '', text)
    
    # 2. 全角数字・ハイフンを半角に
    trans_map = str.maketrans({
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
        '－': '-', 'ー': '-', '−': '-'
    })
    text = text.translate(trans_map)
    
    # 3. 漢数字を算用数字に（簡易）
    kanji_map = str.maketrans({
        '〇': '0', '一': '1', '二': '2', '三': '3', '四': '4',
        '五': '5', '六': '6', '七': '7', '八': '8', '九': '9'
    })
    text = text.translate(kanji_map)
    
    # 4. 住所の区切り文字をハイフンに統一
    text = re.sub(r'(丁目|番地|番|号|の)', '-', text)
    
    # 5. スペース削除・連続ハイフン整理
    text = text.replace(' ', '').replace('　', '').replace('\n', '').replace('\r', '')
    text = re.sub(r'-+', '-', text)
    text = text.strip('-')
    
    return text


def normalize_phone(phone):
    """電話番号を正規化"""
    if not phone or pd.isna(phone):
        return ""
    
    phone = str(phone).strip()
    
    # 全角数字・ハイフンを半角に
    trans_map = str.maketrans({
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
        '－': '-', 'ー': '-', '−': '-'
    })
    phone = phone.translate(trans_map)
    
    # 記号削除
    phone = re.sub(r'[ \-\(\)]', '', phone)
    
    # 国際番号対応 (+81)
    if phone.startswith('+81'):
        phone = '0' + phone[3:]
        
    return phone


def compare_phone(existing_val, new_val):
    """電話番号を比較"""
    norm1 = normalize_phone(existing_val)
    norm2 = normalize_phone(new_val)
    
    if not norm1 and not norm2:
        return False
    
    if not norm1 and norm2:
        return True # 新規取得
        
    return norm1 != norm2


def compare_data(existing_value, new_value, threshold=0.85):
    """既存データと新しいデータを比較（正規化＋類似度判定）"""
    existing_normalized = normalize_text(existing_value)
    new_normalized = normalize_text(new_value)

    # 両方とも空の場合は差異なし
    if not existing_normalized and not new_normalized:
        return False
    
    # 既存が空で新しい値がある場合（新規）→ 差異あり
    if not existing_normalized and new_normalized:
        return True
        
    # 既存があり、新しい値が空（API失敗など）→ 差異ありとするか？
    # 今回の要件では「正しい住所を入れる」なので、Trueとしておく
    if existing_normalized and not new_normalized:
        return True

    # 正規化後の完全一致チェック
    if existing_normalized == new_normalized:
        return False

    # 類似度判定 (difflib)
    ratio = difflib.SequenceMatcher(None, existing_normalized, new_normalized).ratio()
    
    # 類似度が高ければ「同じ（差異なし）」とみなす
    if ratio >= threshold:
        return False
        
    # 類似度が低ければ「差異あり」
    return True


def process_excel(file_path, output_path, max_workers=10, filter_business_type=None):
    """Excelファイルを処理して住所・電話番号などの詳細情報を追加（並列処理版）"""
    try:
        # Excelファイルをpandasで読み込んで新しいワークブックに変換
        logger.info(f"Excelファイル読み込み開始: {file_path}")

        if filter_business_type:
            logger.info(f"業態名フィルタ: {filter_business_type}")

        # まずopenpyxlでヘッダー行を検出
        temp_wb = load_workbook(file_path, data_only=True)
        temp_ws = temp_wb.active

        # ヘッダー行を検出（業態名・店舗名を含む行を探す）
        header_row_idx = 0  # 0-indexed for pandas
        for row in range(1, min(51, temp_ws.max_row + 1)):
            row_values = []
            for col in range(1, temp_ws.max_column + 1):
                value = temp_ws.cell(row=row, column=col).value
                if value:
                    row_values.append(str(value).strip().lower())

            row_str = ' '.join(row_values)
            if '業態' in row_str and ('店舗' in row_str or '店名' in row_str):
                header_row_idx = row - 1  # pandas is 0-indexed
                logger.info(f"ヘッダー行を検出: 行{row} (pandas index: {header_row_idx})")
                break

        # pandasでデータを読み込む（検出したヘッダー行を使用）
        df = pd.read_excel(file_path, engine='openpyxl', header=header_row_idx)
        logger.info(f"pandas読み込み成功: {len(df)}行 x {len(df.columns)}列")
        logger.info(f"列名: {df.columns.tolist()}")

        # pandasのDataFrameから業態名と店舗名の列を特定
        business_type_col_idx = None # 0-based index from df
        store_name_col_idx = None    # 0-based index from df
        original_address_col_idx = None # 0-based index from df
        original_phone_col_idx = None # 0-based index from df
        
        # 業態名列の特定
        for idx, col_name in enumerate(df.columns):
            col_str = str(col_name).strip()
            if col_str == '業態名' or col_str == '業態':
                business_type_col_idx = idx
            elif '業態' in col_str and business_type_col_idx is None:
                business_type_col_idx = idx
            
            # 住所列の特定
            if '住所' in col_str or 'address' in col_str.lower():
                if original_address_col_idx is None:
                    original_address_col_idx = idx
                elif col_str == '住所': # 完全一致を優先
                    original_address_col_idx = idx
                    
            # 電話番号列の特定
            if '電話' in col_str or 'tel' in col_str or 'phone' in col_str:
                if original_phone_col_idx is None:
                    original_phone_col_idx = idx
                elif '電話' in col_str: # 完全一致に近いものを優先
                    original_phone_col_idx = idx
        
        # 店舗名列の特定（優先度付き）
        # 1. 完全一致
        for idx, col_name in enumerate(df.columns):
            col_str = str(col_name).strip()
            if col_str in ['店舗名', '店名', '支店名', '店舗', '支店']:
                store_name_col_idx = idx
                break
        
        # 2. 部分一致（ただし「店舗清掃」などは除外）
        if store_name_col_idx is None:
            for idx, col_name in enumerate(df.columns):
                col_str = str(col_name).strip()
                # 除外キーワード
                if '清掃' in col_str or '時間' in col_str or '日' in col_str or 'No' in col_str:
                    continue
                    
                if '店舗' in col_str or '店名' in col_str or '支店' in col_str:
                    store_name_col_idx = idx
                    break

        if business_type_col_idx is None and store_name_col_idx is None:
            store_name_col_idx = 0
            logger.info(f"業態名・店舗名の列が見つかりません。1列目を検索キーワードとして使用します")

        # 新しい列構成の構築
        new_columns = []
        target_columns = {} # key -> new_col_idx (1-based)
        original_col_map = {} # old_col_idx (0-based) -> new_col_idx (1-based)
        
        current_col = 1
        
        # 既存列を走査しながら新しい列リストを作成
        for i, col_name in enumerate(df.columns):
            col_str = str(col_name).strip()
            col_lower = col_str.lower()
            
            # 既存の「結果カラム」っぽいものはスキップ（重複防止＆位置修正のため）
            # これらは後で適切な位置（住所や電話の隣）に挿入される
            if col_str in ['住所(検索結果)', '電話番号検索結果', '電話番号(検索結果)']:
                continue

            # 現在の列を新しいリストに追加
            new_columns.append(col_name)
            original_col_map[i] = current_col
            
            # --- カラムの役割判定 ---
            is_address_col = False
            is_phone_col = False
            
            # 住所カラムの判定
            if col_str == '住所':
                target_columns['address_orig'] = current_col
                is_address_col = True
            elif ('住所' in col_lower or 'address' in col_lower) and 'address_orig' not in target_columns:
                target_columns['address_orig'] = current_col
                is_address_col = True
            
            # 電話番号カラムの判定
            if col_str == '電話番号' or col_str == '電話':
                target_columns['phone_orig'] = current_col
                is_phone_col = True
            elif ('電話' in col_lower or 'phone' in col_lower) and 'phone_orig' not in target_columns:
                target_columns['phone_orig'] = current_col
                is_phone_col = True
                
            # 営業時間カラムの判定
            if ('営業時間' in col_lower or '営業' in col_lower) and 'business_hours' not in target_columns:
                target_columns['business_hours'] = current_col

            current_col += 1
            
            # --- 検索結果カラムの挿入 ---
            
            # 住所カラムが見つかったら、その直後に「住所(検索結果)」を挿入
            if is_address_col and 'address_result' not in target_columns:
                new_columns.append('住所(検索結果)')
                target_columns['address_result'] = current_col
                current_col += 1
                
            # 電話番号カラムが見つかったら、その直後に「電話番号検索結果」を挿入
            if is_phone_col and 'phone_result' not in target_columns:
                new_columns.append('電話番号検索結果')
                target_columns['phone_result'] = current_col
                current_col += 1

        # 元のファイルに住所や電話番号がなかった場合、末尾に追加
        if 'address_orig' not in target_columns:
            new_columns.append('住所')
            target_columns['address_orig'] = current_col
            current_col += 1
            
            new_columns.append('住所(検索結果)')
            target_columns['address_result'] = current_col
            current_col += 1
            
        elif 'address_result' not in target_columns:
            # 住所はあるが結果カラムがまだない場合（変なケースだが念のため）
            new_columns.append('住所(検索結果)')
            target_columns['address_result'] = current_col
            current_col += 1
            
        if 'phone_orig' not in target_columns:
            new_columns.append('電話番号')
            target_columns['phone_orig'] = current_col
            current_col += 1
            
            new_columns.append('電話番号検索結果')
            target_columns['phone_result'] = current_col
            current_col += 1

        elif 'phone_result' not in target_columns:
            # 電話はあるが結果カラムがまだない場合
            new_columns.append('電話番号検索結果')
            target_columns['phone_result'] = current_col
            current_col += 1
            
        if 'business_hours' not in target_columns:
            new_columns.append('営業時間')
            target_columns['business_hours'] = current_col
            current_col += 1
            
        # 正式名称は常に末尾に追加
        new_columns.append('正式名称')
        target_columns['building_name'] = current_col
        current_col += 1

        # 新しいワークブックを作成
        wb = Workbook()
        ws = wb.active

        # ヘッダーを書き込む
        for col_idx, col_name in enumerate(new_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # データを書き込む（マッピングに基づいて配置）
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data): # col_idx is 0-based
                if col_idx in original_col_map: # スキップされた列でなければ
                    ws.cell(row=row_idx, column=original_col_map[col_idx], value=value)

        logger.info(f"データ変換完了: {len(df)}行")
        
        # 検索用カラムのインデックスを新しい配置に合わせて更新 (1-based)
        business_type_col = original_col_map[business_type_col_idx] if business_type_col_idx is not None else None
        store_name_col = original_col_map[store_name_col_idx] if store_name_col_idx is not None else None
        
        logger.info(f"使用する列 (新配置) - 業態名: {business_type_col}, 店舗名: {store_name_col}")
        logger.info(f"書き込み先列: {target_columns}")

        # タスクを準備（並列処理用）
        tasks = []
        row_mapping = {}  # インデックスと実際の行番号のマッピング
        skipped_count = 0  # フィルタでスキップされた行数

        for row in range(2, ws.max_row + 1):
            # 業態名と店舗名を組み合わせて検索キーワードを作成
            if business_type_col and store_name_col:
                business_type = ws.cell(row=row, column=business_type_col).value
                store_name = ws.cell(row=row, column=store_name_col).value
            else:
                business_type = None
                store_name = ws.cell(row=row, column=store_name_col).value

            # 空行はスキップ
            if not business_type and not store_name:
                continue

            # 業態名フィルタが指定されている場合は、一致する行だけを処理
            if filter_business_type:
                if not business_type or filter_business_type.lower() not in str(business_type).strip().lower():
                    skipped_count += 1
                    continue

            if (business_type or store_name) and str(store_name or business_type).strip() and str(store_name or business_type) != 'nan':
                idx = len(tasks)
                tasks.append((idx, business_type, store_name))
                row_mapping[idx] = row

        if filter_business_type:
            logger.info(f"フィルタ適用: {skipped_count}件をスキップ、{len(tasks)}件を処理対象としました")

        processed_count = 0
        success_count = 0
        total = len(tasks)
        
        # ハイライト用のスタイル
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        logger.info(f"並列処理開始: {total}件を{max_workers}スレッドで処理")

        # 並列処理で住所を検索
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # タスクを投入
            futures = {
                executor.submit(search_single_row_excel, idx, business_type, store_name, total): idx
                for idx, business_type, store_name in tasks
            }

            # 結果を取得
            for future in as_completed(futures):
                try:
                    idx, data, is_success = future.result()

                    if data is not None:
                        row = row_mapping[idx]

                        # 取得したデータを各列に書き込む
                        if 'building_name' in target_columns:
                            cell = ws.cell(row=row, column=target_columns['building_name'], value=data['building_name'])
                            cell.alignment = Alignment(wrap_text=True)
                        
                        if 'address' in target_columns:
                            # 住所書き込みと比較ハイライト
                            cell = ws.cell(row=row, column=target_columns['address'])
                            cell.alignment = Alignment(wrap_text=True)
                            
                            # 元の住所を取得（比較用）
                        
                        # 住所の処理
                        if is_success:
                            # 検索結果を書き込む
                            if 'address_result' in target_columns:
                                cell_res = ws.cell(row=row, column=target_columns['address_result'], value=data['address'])
                                cell_res.alignment = Alignment(wrap_text=True)
                            
                            # 元の住所を訂正・ハイライト
                            if 'address_orig' in target_columns:
                                cell_orig = ws.cell(row=row, column=target_columns['address_orig'])
                                original_val = cell_orig.value
                                cell_orig.alignment = Alignment(wrap_text=True)
                                
                                # 比較して差異があれば訂正してハイライト
                                if compare_data(original_val, data['address']):
                                    cell_orig.value = data['address']
                                    cell_orig.fill = yellow_fill
                        
                        # 電話番号の処理
                        if is_success:
                            # 検索結果を書き込む
                            if 'phone_result' in target_columns:
                                cell_res = ws.cell(row=row, column=target_columns['phone_result'], value=data['phone'])
                                cell_res.alignment = Alignment(wrap_text=True)
                            
                            # 元の電話番号を訂正・ハイライト
                            if 'phone_orig' in target_columns:
                                cell_orig = ws.cell(row=row, column=target_columns['phone_orig'])
                                original_val = cell_orig.value
                                cell_orig.alignment = Alignment(wrap_text=True)
                                
                                # 比較して差異があれば訂正してハイライト
                                if compare_phone(original_val, data['phone']):
                                    cell_orig.value = data['phone']
                                    cell_orig.fill = yellow_fill
                            
                        if 'business_hours' in target_columns:
                            cell = ws.cell(row=row, column=target_columns['business_hours'], value=data['business_hours'])
                            cell.alignment = Alignment(wrap_text=True)

                        processed_count += 1

                        if is_success:
                            success_count += 1

                        # 進捗表示
                        if processed_count % 10 == 0:
                            logger.info(f"進捗: {processed_count}/{total} 完了 ({success_count}件成功)")

                except Exception as e:
                    logger.error(f"タスク処理エラー: {str(e)}")

        # ファイルを保存
        try:
            # 新しいワークブックとして保存（互換性の問題を回避）
            new_wb = Workbook()
            new_ws = new_wb.active

            # データをコピー
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
                for col_idx, cell in enumerate(row, 1):
                    new_cell = new_ws.cell(row=row_idx, column=col_idx)
                    new_cell.value = cell.value
                    # フォーマットもコピー（基本的なもののみ）
                    if cell.has_style:
                        try:
                            # フォントや罫線だけでなく、塗りつぶしもコピーする
                            new_cell.font = cell.font.copy()
                            new_cell.border = cell.border.copy()
                            if cell.fill and cell.fill.fill_type:
                                new_cell.fill = cell.fill.copy()
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = cell.alignment.copy()
                        except:
                            pass  # スタイルのコピーに失敗しても続行

            new_wb.save(output_path)
            logger.info(f"処理完了: {processed_count}件中{success_count}件の情報を取得しました")

        except Exception as save_error:
            logger.error(f"ファイル保存エラー: {str(save_error)}")
            # 保存に失敗した場合は元の方法で試す
            wb.save(output_path)
            logger.info(f"処理完了（代替保存方法使用）: {processed_count}件中{success_count}件の情報を取得しました")

        # 結果メッセージを作成
        result_message = f"{processed_count}件中{success_count}件の情報を取得しました"
        if filter_business_type:
            result_message = f"フィルタ「{filter_business_type}」適用: " + result_message

        return True, result_message

    except Exception as e:
        logger.error(f"Excel処理エラー: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, f"エラー: {str(e)}"


@app.route('/')
def index():
    """メインページ"""
    api_key_loaded = GOOGLE_MAPS_API_KEY is not None
    return render_template('index_simple.html', api_key_loaded=api_key_loaded)


@app.route('/upload', methods=['POST'])
def upload_file():
    """ファイルアップロードと処理"""
    try:
        # APIキーのチェック
        if not GOOGLE_MAPS_API_KEY:
            flash('Google Maps APIキーが設定されていません。config.txtを確認してください', 'error')
            return redirect(url_for('index'))

        # 業態名フィルタの取得（オプション）
        filter_business_type = request.form.get('filter_business_type', '').strip()

        # ファイルの取得
        if 'file' not in request.files:
            flash('ファイルが選択されていません', 'error')
            return redirect(url_for('index'))

        file = request.files['file']
        if file.filename == '':
            flash('ファイルが選択されていません', 'error')
            return redirect(url_for('index'))

        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash('Excelファイル(.xlsx または .xls)を選択してください', 'error')
            return redirect(url_for('index'))

        # ファイルを保存
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(upload_path)

        # ファイル形式の検証と変換
        is_valid, result = validate_excel_file(upload_path)
        if not is_valid:
            flash(f'ファイル形式エラー: {result}', 'error')
            return redirect(url_for('index'))

        # 変換されたファイルがある場合はそれを使用
        if isinstance(result, str) and result.endswith('_converted.xlsx'):
            upload_path = result
            filename = os.path.basename(result)
            flash('ファイル形式を自動変換しました', 'success')

        # 出力ファイルパス
        output_filename = f"住所追加_{filename}".replace('_converted.xlsx', '.xlsx')
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

        # Excelファイルを処理（並列処理版）
        success, message = process_excel(upload_path, output_path, max_workers=10, filter_business_type=filter_business_type)

        if success:
            flash(message, 'success')
            return send_file(
                output_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash(message, 'error')
            return redirect(url_for('index'))

    except Exception as e:
        logger.error(f"アップロードエラー: {str(e)}")
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('index'))


if __name__ == '__main__':
    # APIキーを読み込む
    load_api_key()

    if not GOOGLE_MAPS_API_KEY:
        print("\n" + "=" * 60)
        print("⚠ 警告: Google Maps APIキーが設定されていません")
        print("=" * 60)
        print("config.txt ファイルを作成して以下の形式でAPIキーを設定してください:")
        print("GOOGLE_MAPS_API_KEY=あなたのAPIキー")
        print("=" * 60)
        print("\nAPIキーの取得方法は GOOGLE_MAPS_API_SETUP.md を参照してください")
        print("=" * 60)
        print("\nアプリは起動しますが、住所検索は動作しません\n")

    app.run(debug=True, host='0.0.0.0', port=5001)
